"""桩基工序照片批量统计与完整性初筛工具。

本工具只根据文件名判断某个桩号是否收集到对应工序的照片，
不读取照片内容，也不判断施工质量或验收结论。
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp"}
STANDARD_PROCESSES = ("对中", "焊接", "钢筋笼验收", "初灌", "二清", "入岩", "终孔")
PROCESS_ALIASES = {"钢筋笼焊接": "焊接"}
FILENAME_PATTERN = re.compile(r"^([A-Za-z0-9]+)([\u4e00-\u9fff].*)$")


@dataclass(frozen=True)
class PhotoRecord:
    """一张候选照片的必要信息。"""

    date_folder: str
    pile_folder: str
    filename: str
    relative_path: str


@dataclass
class PileGroup:
    """同一基础桩号下识别到的桩号与工序集合。"""

    pile_nos: set[str] = field(default_factory=set)
    processes: set[str] = field(default_factory=set)


@dataclass
class AnalysisResult:
    """一次扫描的结构化结果。"""

    total_photos: int
    groups: dict[str, PileGroup]
    skipped_files: list[PhotoRecord]

    @property
    def parsed_photos(self) -> int:
        return self.total_photos - len(self.skipped_files)


def collect_photo_info(root_folder: Path | str) -> list[PhotoRecord]:
    """按“日期文件夹/桩号文件夹/照片”结构收集候选图片。"""

    root = Path(root_folder).expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"输入目录不存在或不是文件夹：{root}")

    photo_data: list[PhotoRecord] = []
    for date_path in sorted(path for path in root.iterdir() if path.is_dir()):
        for pile_path in sorted(path for path in date_path.iterdir() if path.is_dir()):
            for file_path in sorted(path for path in pile_path.iterdir() if path.is_file()):
                if file_path.suffix.lower() not in IMAGE_SUFFIXES:
                    continue
                photo_data.append(
                    PhotoRecord(
                        date_folder=date_path.name,
                        pile_folder=pile_path.name,
                        filename=file_path.name,
                        relative_path=file_path.relative_to(root).as_posix(),
                    )
                )
    return photo_data


def extract_pile_info(filename: str) -> tuple[str | None, str | None]:
    """从文件名提取桩号和工序。

    规则：文件名以字母或数字开头；首个汉字左侧为桩号，首个汉字及其右侧为工序。
    """

    name_without_ext = Path(filename).stem.strip()
    match = FILENAME_PATTERN.match(name_without_ext)
    if not match:
        return None, None
    return match.group(1), match.group(2).strip()


def normalize_process(process_name: str) -> str:
    """将工序别名和带编号变体归一为标准工序，其余归为“其它”。"""

    cleaned_name = process_name.strip()
    normalized_name = PROCESS_ALIASES.get(cleaned_name, cleaned_name)
    if normalized_name in STANDARD_PROCESSES:
        return normalized_name
    for standard_process in STANDARD_PROCESSES:
        if standard_process in normalized_name:
            return standard_process
    return "其它"


def get_pile_group_key(pile_no: str) -> str:
    """按首个大写字母 D 左侧内容合并带后缀桩号。"""

    prefix, separator, _ = pile_no.partition("D")
    return prefix if separator and prefix else pile_no


def analyze_process_completeness(photo_data: Iterable[PhotoRecord]) -> AnalysisResult:
    """解析照片文件名，形成“基础桩号 -> 已收集工序”的统计结果。"""

    photos = list(photo_data)
    groups: defaultdict[str, PileGroup] = defaultdict(PileGroup)
    skipped_files: list[PhotoRecord] = []

    for photo in photos:
        pile_no, process = extract_pile_info(photo.filename)
        if pile_no is None or process is None:
            skipped_files.append(photo)
            continue

        group_key = get_pile_group_key(pile_no)
        groups[group_key].pile_nos.add(pile_no)
        groups[group_key].processes.add(normalize_process(process))

    return AnalysisResult(
        total_photos=len(photos),
        groups=dict(groups),
        skipped_files=skipped_files,
    )


def _missing_processes(processes: set[str]) -> list[str]:
    return [process for process in STANDARD_PROCESSES if process not in processes]


def generate_text_report(result: AnalysisResult, output_file: Path | str) -> Path:
    """生成便于快速查看的 UTF-8 文本报告。"""

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "桩基工序照片完整性初筛报告",
        "=" * 42,
        "说明：结果仅表示是否检索到符合命名规则的照片，不代表工序完成或验收合格。",
        f"扫描照片：{result.total_photos}",
        f"成功解析：{result.parsed_photos}",
        f"命名异常：{len(result.skipped_files)}",
        f"桩号分组：{len(result.groups)}",
        "",
    ]

    for group_key in sorted(result.groups):
        group = result.groups[group_key]
        existing = [process for process in STANDARD_PROCESSES if process in group.processes]
        if "其它" in group.processes:
            existing.append("其它")
        missing = _missing_processes(group.processes)
        lines.extend(
            [
                f"桩号分组：{group_key}",
                f"  包含桩号：{', '.join(sorted(group.pile_nos))}",
                f"  已有工序：{', '.join(existing) if existing else '无'}",
                f"  待核查项：{', '.join(missing) if missing else '无'}",
                "",
            ]
        )

    if result.skipped_files:
        lines.append("命名异常文件：")
        lines.extend(f"  - {photo.relative_path}" for photo in result.skipped_files)

    output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return output_path


def generate_excel_report(result: AnalysisResult, output_file: Path | str) -> Path:
    """生成带筛选、冻结窗格和状态配色的 Excel 核查矩阵。"""

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "核查矩阵"
    matrix.sheet_view.showGridLines = False

    headers = [
        "桩号分组",
        "包含桩号",
        *STANDARD_PROCESSES,
        "其它",
        "已收集标准工序数",
        "待核查数",
        "已有工序汇总",
        "待核查项",
    ]
    last_column = len(headers)
    last_column_letter = matrix.cell(row=1, column=last_column).column_letter

    matrix.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    matrix["A1"] = "桩基工序照片完整性初筛矩阵"
    matrix["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    matrix["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    matrix["A1"].alignment = Alignment(horizontal="center", vertical="center")
    matrix.row_dimensions[1].height = 30

    matrix.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    matrix["A2"] = (
        f"共扫描 {result.total_photos} 张照片，成功解析 {result.parsed_photos} 张，"
        f"形成 {len(result.groups)} 个桩号分组；结果仅供资料完整性初筛。"
    )
    matrix["A2"].font = Font(name="Microsoft YaHei", size=10, color="404040")
    matrix["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    matrix["A2"].alignment = Alignment(horizontal="left", vertical="center")
    matrix.row_dimensions[2].height = 24

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9E2F3")
    for column, header in enumerate(headers, start=1):
        cell = matrix.cell(row=3, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
    matrix.row_dimensions[3].height = 34

    for row_index, group_key in enumerate(sorted(result.groups), start=4):
        group = result.groups[group_key]
        existing_ordered = [process for process in STANDARD_PROCESSES if process in group.processes]
        if "其它" in group.processes:
            existing_ordered.append("其它")
        missing = _missing_processes(group.processes)
        values = [
            group_key,
            ", ".join(sorted(group.pile_nos)),
            *("✓" if process in group.processes else "✗" for process in STANDARD_PROCESSES),
            "✓" if "其它" in group.processes else "✗",
            len(existing_ordered) - (1 if "其它" in group.processes else 0),
            len(missing),
            " | ".join(existing_ordered),
            " | ".join(missing) if missing else "无",
        ]
        for column, value in enumerate(values, start=1):
            cell = matrix.cell(row=row_index, column=column, value=value)
            cell.font = Font(name="Microsoft YaHei", size=10, color="262626")
            cell.alignment = Alignment(
                horizontal="center" if column <= 12 else "left",
                vertical="center",
                wrap_text=column >= 13,
            )
            cell.border = Border(bottom=thin_gray)
        if row_index % 2 == 0:
            for column in range(1, last_column + 1):
                matrix.cell(row=row_index, column=column).fill = PatternFill("solid", fgColor="F7FAFC")

    last_row = max(4, 3 + len(result.groups))
    status_range = f"C4:J{last_row}"
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    green_font = Font(color="375623", bold=True)
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    red_font = Font(color="C00000", bold=True)
    matrix.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"✓"'], fill=green_fill, font=green_font),
    )
    matrix.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"✗"'], fill=red_fill, font=red_font),
    )

    matrix.freeze_panes = "C4"
    matrix.auto_filter.ref = f"A3:{last_column_letter}{last_row}"
    widths = {
        "A": 13,
        "B": 22,
        "C": 10,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 17,
        "L": 12,
        "M": 42,
        "N": 38,
    }
    for column, width in widths.items():
        matrix.column_dimensions[column].width = width
    matrix.print_title_rows = "1:3"
    matrix.page_setup.orientation = "landscape"
    matrix.page_setup.fitToWidth = 1
    matrix.sheet_properties.pageSetUpPr.fitToPage = True

    exceptions = workbook.create_sheet("命名异常")
    exceptions.sheet_view.showGridLines = False
    exceptions.append(["相对路径", "文件名", "处理结果"])
    for cell in exceptions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if result.skipped_files:
        for photo in result.skipped_files:
            exceptions.append([photo.relative_path, photo.filename, "跳过：不符合文件名解析规则"])
    else:
        exceptions.append(["-", "-", "未发现命名异常"])
    exceptions.freeze_panes = "A2"
    exceptions.auto_filter.ref = f"A1:C{exceptions.max_row}"
    exceptions.column_dimensions["A"].width = 48
    exceptions.column_dimensions["B"].width = 28
    exceptions.column_dimensions["C"].width = 34
    for row in exceptions.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    notes = workbook.create_sheet("使用说明")
    notes.sheet_view.showGridLines = False
    notes["A1"] = "结果边界"
    notes["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    notes["A2"] = "1. 本工具只根据文件名判断照片是否已收集。"
    notes["A3"] = "2. “✗”表示未检索到符合规则的照片，需要人工复核，不等于工序未施工。"
    notes["A4"] = "3. 工具不识别照片画面内容，不判断施工质量，也不替代验收程序。"
    notes["A5"] = "4. 公开仓库示例全部为合成数据，不含真实工程资料。"
    for row in range(2, 6):
        notes[f"A{row}"].font = Font(name="Microsoft YaHei", size=11, color="404040")
        notes[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 88
    notes.row_dimensions[1].height = 30
    for row in range(2, 6):
        notes.row_dimensions[row].height = 26

    workbook.properties.creator = "Pile Photo Completeness Checker"
    workbook.properties.lastModifiedBy = "Pile Photo Completeness Checker"
    workbook.properties.title = "桩基工序照片完整性初筛演示"
    workbook.properties.subject = "Synthetic portfolio demo"
    workbook.properties.description = "All rows are synthetic demo data."
    workbook.save(output_path)
    return output_path


def run_analysis(input_dir: Path | str, output_dir: Path | str) -> tuple[AnalysisResult, Path, Path]:
    """完成扫描、分析并生成 TXT 与 XLSX 两种结果。"""

    photos = collect_photo_info(input_dir)
    result = analyze_process_completeness(photos)
    output_root = Path(output_dir)
    text_path = generate_text_report(result, output_root / "pile_photo_completeness.txt")
    excel_path = generate_excel_report(result, output_root / "pile_photo_completeness.xlsx")
    return result, text_path, excel_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="按照片文件名生成“桩号 × 工序”完整性初筛矩阵。"
    )
    parser.add_argument(
        "input_dir",
        nargs="?",
        default="examples/demo_input",
        help="输入目录，结构为：日期文件夹/桩号文件夹/照片（默认：examples/demo_input）",
    )
    parser.add_argument(
        "--output",
        default="examples/demo_output",
        help="结果输出目录（默认：examples/demo_output）",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        result, text_path, excel_path = run_analysis(args.input_dir, args.output)
    except (FileNotFoundError, PermissionError, OSError) as error:
        print(f"处理失败：{error}")
        return 1

    print("统计完成")
    print(f"- 扫描照片：{result.total_photos}")
    print(f"- 成功解析：{result.parsed_photos}")
    print(f"- 命名异常：{len(result.skipped_files)}")
    print(f"- 桩号分组：{len(result.groups)}")
    print(f"- 文本报告：{text_path.resolve()}")
    print(f"- Excel 矩阵：{excel_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
