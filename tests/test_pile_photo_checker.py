from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pile_photo_checker import (
    PhotoRecord,
    analyze_process_completeness,
    extract_pile_info,
    get_pile_group_key,
    normalize_process,
    run_analysis,
)


class RuleTests(unittest.TestCase):
    def test_extracts_pile_and_process(self) -> None:
        self.assertEqual(extract_pile_info("P018钢筋笼验收.jpg"), ("P018", "钢筋笼验收"))

    def test_rejects_filename_without_alphanumeric_prefix(self) -> None:
        self.assertEqual(extract_pile_info("现场照片.jpg"), (None, None))

    def test_normalizes_alias_numbered_variant_and_other(self) -> None:
        self.assertEqual(normalize_process("钢筋笼焊接"), "焊接")
        self.assertEqual(normalize_process("焊接1"), "焊接")
        self.assertEqual(normalize_process("旁站记录"), "其它")

    def test_merges_suffix_after_uppercase_d(self) -> None:
        self.assertEqual(get_pile_group_key("P003D900"), "P003")
        self.assertEqual(get_pile_group_key("P003"), "P003")

    def test_analysis_merges_groups_and_tracks_invalid_names(self) -> None:
        photos = [
            PhotoRecord("2025-10-20", "P003", "P003对中.jpg", "a/P003对中.jpg"),
            PhotoRecord("2025-10-20", "P003", "P003D900终孔.jpg", "a/P003D900终孔.jpg"),
            PhotoRecord("2025-10-20", "P003", "现场照片.jpg", "a/现场照片.jpg"),
        ]
        result = analyze_process_completeness(photos)
        self.assertEqual(result.total_photos, 3)
        self.assertEqual(result.parsed_photos, 2)
        self.assertEqual(len(result.skipped_files), 1)
        self.assertEqual(set(result.groups), {"P003"})
        self.assertEqual(result.groups["P003"].pile_nos, {"P003", "P003D900"})
        self.assertEqual(result.groups["P003"].processes, {"对中", "终孔"})


class IntegrationTests(unittest.TestCase):
    def test_generates_text_and_excel_reports(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "input" / "2025-10-20" / "P001"
            input_dir.mkdir(parents=True)
            for filename in ("P001对中.jpg", "P001焊接.jpg", "现场照片.jpg"):
                (input_dir / filename).write_bytes(b"")

            result, text_path, excel_path = run_analysis(root / "input", root / "output")

            self.assertEqual(result.total_photos, 3)
            self.assertEqual(result.parsed_photos, 2)
            self.assertTrue(text_path.is_file())
            self.assertTrue(excel_path.is_file())

            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            self.assertEqual(workbook.sheetnames, ["核查矩阵", "命名异常", "使用说明"])
            self.assertEqual(workbook["核查矩阵"]["A4"].value, "P001")
            self.assertEqual(workbook["命名异常"]["B2"].value, "现场照片.jpg")


if __name__ == "__main__":
    unittest.main()
